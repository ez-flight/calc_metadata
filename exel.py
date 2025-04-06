# Импорт необходимых библиотек
import math  # Математические функции
from datetime import datetime, timedelta  # Работа с датой и временем
from typing import Any, Dict, List, Optional, Tuple  # Аннотации типов

import shapefile  # Создание/чтение ESRI shapefiles
from openpyxl import Workbook  # Работа с Excel файлами
from openpyxl.utils import \
    get_column_letter  # Вспомогательные функции для Excel
from pyorbital.orbital import Orbital  # Расчет орбитальной механики

# Локальные модули
from calc_cord import get_xyzv_from_latlon  # Конвертация координат
from calc_F_L import (calc_f_doplera,  # Расчет доплеровских параметров
                      calc_lamda)
from read_TBF import (read_tle_base_file,  # Чтение TLE данных
                      read_tle_base_internet)

# ----------------------------
# Глобальные константы
# ----------------------------
EARTH_RADIUS = 6378.140  # Радиус Земли в километрах
EARTH_ROTATION_RATE = 7.2292115E-5  # Угловая скорость вращения Земли (рад/с)
WAVELENGTH = 0.000096  # Длина волны зондирующего сигнала в метрах
EARTH_ANGULAR_VELOCITY = 1674  # Угловая скорость Земли (км/с)
ORBITAL_PERIOD = 5689  # Период обращения спутника в секундах

# ----------------------------
# Класс для хранения параметров наблюдения
# ----------------------------
class ObservationParameters:
    """Контейнер для параметров съемки"""
    def __init__(self):
        self.dt_start = None  # Начальное время наблюдения
        self.dt_end = None    # Конечное время наблюдения
        self.delta = None     # Шаг общего расчета (сек)
        self.delta_obn = None # Шаг расчета во время съемки (сек)
        self.pos_gt = None    # Координаты наземной точки (широта, долгота, высота)

# ----------------------------
# Класс для работы с орбитой спутника
# ----------------------------
class SatelliteTracker:
    """Орбитальный калькулятор на основе TLE"""
    def __init__(self, tle_1: str, tle_2: str):
        # Инициализация библиотеки pyorbital
        self.orb = Orbital("N", line1=tle_1, line2=tle_2)
        self.tle_1 = tle_1  # Первая строка TLE
        self.tle_2 = tle_2  # Вторая строка TLE
    
    def get_geodetic_position(self, utc_time: datetime) -> Tuple[float, float, float]:
        """Возвращает географические координаты спутника (долгота, широта, высота)"""
        return self.orb.get_lonlatalt(utc_time)
    
    def get_inertial_position(self, utc_time: datetime) -> Tuple[Tuple[float, float, float], Tuple[float, float, float]]:
        """Возвращает позицию и скорость в инерциальной системе координат"""
        return self.orb.get_position(utc_time, False)

# ----------------------------
# Расчет геометрических параметров
# ----------------------------
def calculate_angles(R_s: Tuple[float, float, float], 
                    V_s: Tuple[float, float, float],
                    pos_it: Tuple[float, float, float]) -> Dict[str, float]:
    """
    Рассчитывает углы и расстояния между спутником и наземной точкой
    
    Параметры:
    R_s - координаты спутника в инерциальной системе (X, Y, Z)
    V_s - вектор скорости спутника
    pos_it - координаты наземной точки в инерциальной системе
    
    Возвращает словарь с параметрами:
    R_s, R_e, R_0 - расстояния (спутник-центр Земли, точка-центр, спутник-точка)
    y_grad - угол визирования в градусах
    ay_grad - угол места в градусах
    """
    X_s, Y_s, Z_s = R_s
    X_t, Y_t, Z_t = pos_it
    
    # Расчет расстояний
    R_s_norm = math.sqrt(X_s**2 + Y_s**2 + Z_s**2)  # Норма вектора R_s
    R_e_norm = math.sqrt(X_t**2 + Y_t**2 + Z_t**2)  # Норма вектора R_e
    R_0_norm = math.sqrt((X_s-X_t)**2 + (Y_s-Y_t)**2 + (Z_s-Z_t)**2)  # Расстояние спутник-точка
    V_s_norm = math.sqrt(V_s[0]**2 + V_s[1]**2 + V_s[2]**2)  # Скорость спутника
    
    # Расчет углов по формулам сферической геометрии
    y = math.acos((R_0_norm**2 + R_s_norm**2 - R_e_norm**2) / (2 * R_0_norm * R_s_norm))
    ay = math.acos((R_0_norm * math.sin(y)) / R_e_norm)
    
    return {
        'R_s': R_s_norm,
        'R_e': R_e_norm,
        'R_0': R_0_norm,
        'V_s': V_s_norm,
        'y_grad': math.degrees(y),     # Угол визирования
        'ay_grad': math.degrees(ay)    # Угол места
    }

# ----------------------------
# Построение орбитального трека
# ----------------------------
def create_orbital_track(tracker: SatelliteTracker, 
                        params: ObservationParameters,
                        track_shape: Optional[shapefile.Writer] = None,
                        angle_param: float = 0) -> List[Tuple]:
    """
    Генерирует трек спутника с учетом условий наблюдения
    
    Параметры:
    tracker - экземпляр орбитального трекера
    params - параметры наблюдения
    track_shape - объект для записи в shapefile
    angle_param - параметр угла для расчета доплера
    
    Возвращает список кортежей с данными по виткам
    """
    lat_t, lon_t, alt_t = params.pos_gt  # Координаты наземной станции
    t_semki = []  # Собранные данные съемки
    current_orbit = 0  # Текущий номер витка
    orbit_start_time = None  # Время начала витка
    orbit_points = []  # Точки текущего витка
    
    dt = params.dt_start  # Текущее время расчета
    while dt < params.dt_end:
        # Получение текущих координат спутника
        lon_s, lat_s, alt_s = tracker.get_geodetic_position(dt)
        R_s, V_s = tracker.get_inertial_position(dt)
        pos_it, _ = get_xyzv_from_latlon(dt, lon_t, lat_t, alt_t)  # Конвертация координат станции
        
        angles = calculate_angles(R_s, V_s, pos_it)  # Расчет углов
        
        # Проверка условий для съемки
        if 24 <= angles['y_grad'] <= 55 and angles['R_0'] < angles['R_e']:
            # Расчет доплеровской частоты
            Fd = calc_f_doplera(
                angle_param, WAVELENGTH, math.radians(angles['ay_grad']), 
                R_s, V_s, angles['R_0'], angles['R_e'], 
                angles['R_s'], angles['V_s']
            )
            
            # Определение номера витка
            orbit_num = int((dt - params.dt_start).total_seconds() // ORBITAL_PERIOD) + 1
            
            # Обработка смены витка
            if orbit_num != current_orbit:
                if current_orbit != 0 and orbit_start_time and orbit_points:
                    t_semki.append((
                        current_orbit,
                        orbit_points.copy(),
                        orbit_start_time,
                        dt,
                        dt - orbit_start_time
                    ))
                current_orbit = orbit_num
                orbit_start_time = dt
                orbit_points = []
            
            # Добавление точки съемки
            orbit_points.append((dt, (lon_s, lat_s, alt_s), Fd))
            
            # Запись в shapefile
            if track_shape:
                track_shape.point(lon_s, lat_s)  # Добавление точки
                track_shape.record(  # Запись атрибутов
                    current_orbit,
                    dt.strftime("%Y-%m-%d %H:%M:%S.%f"),  # Время с микросекундами
                    lon_s, lat_s,
                    float(angles['R_s']),
                    float(angles['R_e']),
                    float(angles['R_0']),
                    float(angles['y_grad']),
                    float(angles['ay_grad']),
                    float(WAVELENGTH),
                    float(Fd)
                )
            
            # Увеличение времени с шагом съемки
            dt += params.delta_obn
        else:
            # Увеличение времени с общим шагом
            dt += params.delta
    
    # Добавление последнего витка
    if orbit_start_time and orbit_points:
        t_semki.append((
            current_orbit,
            orbit_points,
            orbit_start_time,
            dt,
            dt - orbit_start_time
        ))
    
    return t_semki

# ----------------------------
# Экспорт данных в Excel
# ----------------------------
def save_to_excel(data: List[Tuple], filename: str, params):
    """Сохраняет рассчитанные данные в Excel файл с автонастройкой ширины столбцов"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Satellite Data"
    
    # Формирование заголовков
    headers = [
        "Виток", "Время", "Долгота", "Широта", "Высота",
        "Fd (Гц)", "R_s (км)", "R_e (км)", "R_0 (км)",
        "Угол визирования (°)", "Угол места (°)"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2  # Начальная строка данных
    for orbit in data:
        orbit_num, points, start_time, end_time, duration = orbit
        for point in points:
            dt, (lon, lat, alt), Fd = point
            # Расчет углов для текущей точки
            angles = calculate_angles(
                tracker.get_inertial_position(dt)[0],
                tracker.get_inertial_position(dt)[1],
                get_xyzv_from_latlon(dt, *params.pos_gt)[0]
            )
            
            # Запись данных в строку
            ws.cell(row=row, column=1, value=orbit_num)
            ws.cell(row=row, column=2, value=dt.strftime("%Y-%m-%d %H:%M:%S.%f"))  # Время с микросекундами
            ws.cell(row=row, column=3, value=lon)
            ws.cell(row=row, column=4, value=lat)
            ws.cell(row=row, column=5, value=alt)
            ws.cell(row=row, column=6, value=Fd)
            ws.cell(row=row, column=7, value=angles['R_s'])
            ws.cell(row=row, column=8, value=angles['R_e'])
            ws.cell(row=row, column=9, value=angles['R_0'])
            ws.cell(row=row, column=10, value=angles['y_grad'])
            ws.cell(row=row, column=11, value=angles['ay_grad'])
            
            row += 1
    
    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filename)
    print(f"Данные сохранены в Excel файл: {filename}")

# ----------------------------
# Тестовая функция
# ----------------------------
def _test():
    """Пример использования функционала"""
    # Инициализация параметров спутника
    s_name, tle_1, tle_2 = read_tle_base_file(56756)  # Чтение TLE из файла
    pos_gt = (59.95, 30.316667, 0)  # Координаты Санкт-Петербурга
    
    # Настройка параметров наблюдения
    params = ObservationParameters()
    params.pos_gt = pos_gt
    params.dt_start = datetime(2024, 2, 21, 3, 0, 0)  # Начало наблюдения
    params.delta = timedelta(seconds=10)  # Основной шаг расчета
    params.delta_obn = timedelta(seconds=5)  # Шаг во время съемки
    params.dt_end = params.dt_start + timedelta(days=2)  # Окончание наблюдения
    
    # Инициализация трекера
    global tracker
    tracker = SatelliteTracker(tle_1, tle_2)
    
    # Настройка shapefile
    shp_filename = f"result/{s_name}"
    track_shape = shapefile.Writer(shp_filename, shapefile.POINT)
    
    try:
        # Определение полей shapefile
        fields = [
            ("ID", "N", 10),          # Номер витка
            ("TIME", "C", 40),        # Время с микросекундами
            ("LON", "F", 10, 10),     # Долгота
            ("LAT", "F", 10, 10),     # Широта
            ("R_s", "F", 10, 5),      # Расстояние спутник-центр Земли
            ("R_t", "F", 10, 5),      # Расстояние точка-центр Земли
            ("R_n", "F", 10, 5),      # Расстояние спутник-точка
            ("ϒ", "F", 10, 5),        # Угол визирования
            ("φ", "F", 10, 5),        # Угол места
            ("λ", "F", 10, 10),       # Длина волны
            ("f", "F", 10, 5)         # Доплеровская частота
        ]
        
        # Добавление полей в shapefile
        for field in fields:
            if len(field) == 3:
                track_shape.field(*field)
            else:
                track_shape.field(field[0], field[1], field[2], field[3])
        
        # Расчет для диапазона углов
        angle_values = range(85, 95, 1)  # Углы от 85° до 94° с шагом 1°
        results = {}
        
        for angle in angle_values:
            # Генерация трека
            t_semki = create_orbital_track(tracker, params, track_shape, angle)
            results[angle] = t_semki
            
            # Экспорт в Excel
            excel_filename = f"result/{s_name}_angle_{angle}.xlsx"
            save_to_excel(t_semki, excel_filename, params)
            
            # Вывод статистики
            print(f"\nУгол {angle}°:")
            for orbit in t_semki:
                fd_values = [point[2] for point in orbit[1]]
                if fd_values:
                    print(f"Виток {orbit[0]}: точек {len(orbit[1])}, Fd от {min(fd_values):.2f} до {max(fd_values):.2f} Гц")
        
        # Финализация shapefile
        track_shape.save(shp_filename)
        print(f"\nShapefile сохранен: {shp_filename}.shp")
        
    except Exception as e:
        print(f"Ошибка: {str(e)}")
    finally:
        # Гарантированное закрытие shapefile
        if 'track_shape' in locals():
            try:
                track_shape.close()
            except:
                pass
    
    # Создание PRJ-файла с проекцией WGS84
    try:
        with open(f"{shp_filename}.prj", "w") as prj:
            prj.write('GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]')
    except Exception as e:
        print(f"Ошибка при создании PRJ-файла: {e}")

# Точка входа в программу
if __name__ == "__main__":
    _test()