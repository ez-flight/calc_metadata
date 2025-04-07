"""
Модуль для расчета орбитальных параметров спутника и условий съемки наземных объектов
"""

# Импорт необходимых библиотек
import math  # Математические функции и константы
from datetime import datetime, timedelta  # Работа с датой и временем
from typing import Any, Dict, List, Optional, Tuple  # Аннотации типов
from concurrent.futures import ThreadPoolExecutor, as_completed
import numpy as np
from multiprocessing import Pool
from dataclasses import dataclass
import shapefile  # Создание и запись GIS shapefiles
from openpyxl import Workbook  # Генерация Excel-отчетов
from openpyxl.utils import \
    get_column_letter  # Вспомогательные функции для Excel
from pyorbital.orbital import Orbital  # Расчет орбитальной механики
from multiprocessing import Pool
# Локальные модули
from calc_cord import \
    get_xyzv_from_latlon  # Конвертация геодезических координат в ECEF
from calc_F_L import calc_f_doplera  # Расчет доплеровских параметров
from calc_F_L import calc_lamda
from read_TBF import read_tle_base_file  # Чтение TLE данных
from read_TBF import read_tle_base_internet

# ----------------------------
# Глобальные константы
# ----------------------------
EARTH_RADIUS = 6378.140  # Экваториальный радиус Земли в километрах
EARTH_ROTATION_RATE = 7.2292115E-5  # Угловая скорость вращения Земли (рад/с)
WAVELENGTH = 0.000096  # Длина волны зондирующего сигнала в метрах (X-диапазон)
ORBITAL_PERIOD = 5689  # Типичный период обращения спутника в секундах (~94 мин)

@dataclass
class OrbitChunk:
    start_time: datetime
    end_time: datetime
    points: List[Tuple]


class ObservationParameters:
    """
    Контейнер для хранения параметров наблюдения.
    
    Атрибуты:
    dt_start (datetime): Начальное время наблюдения
    dt_end (datetime): Конечное время наблюдения
    delta (timedelta): Основной шаг расчета траектории
    delta_obn (timedelta): Шаг расчета во время съемки
    pos_gt (tuple): Координаты наземной станции (широта, долгота, высота)
    """
    def __init__(self):
        self.dt_start = None
        self.dt_end = None
        self.delta = None
        self.delta_obn = None
        self.pos_gt = None

class SatelliteTracker:
    """
    Класс для расчета орбитальных параметров спутника на основе TLE.
    
    Методы:
    - get_geodetic_position: Возвращает географические координаты спутника
    - get_inertial_position: Возвращает позицию и скорость в инерциальной СК
    - check_orbit_transition: Определяет момент смены витка
    - predict_next_entry: Прогнозирует следующее вхождение в зону съемки
    - find_exact_boundary: Точное определение границ зоны съемки
    - calculate_angles: Расчет геометрических параметров
    - is_in_shooting_zone: Проверка условий для проведения съемки
    """
    
    def __init__(self, tle_1: str, tle_2: str):
        """
        Инициализация трекера на основе двухстрочного TLE.
        
        Параметры:
        tle_1 (str): Первая строка TLE
        tle_2 (str): Вторая строка TLE
        """
        self.orb = Orbital("N", line1=tle_1, line2=tle_2)
        self.tle_1 = tle_1
        self.tle_2 = tle_2
        self._prev_lat = None  # Для отслеживания пересечения экватора

    def get_geodetic_position(self, utc_time: datetime) -> Tuple[float, float, float]:
        """Возвращает (долготу, широту, высоту) спутника в заданное время."""
        return self.orb.get_lonlatalt(utc_time)

    def get_inertial_position(self, utc_time: datetime) -> Tuple[Tuple[float, float, float], Tuple[float, float, float]]:
        """Возвращает позицию (X,Y,Z) и скорость (Vx,Vy,Vz) в инерциальной системе координат."""
        return self.orb.get_position(utc_time, False)

    def check_orbit_transition(self, dt: datetime) -> bool:
        """
        Определяет момент смены витка по пересечению экватора.
        
        Логика: переход с южного полушария на северное (широта меняется с отрицательной на положительную)
        """
        _, lat, _ = self.get_geodetic_position(dt)
        if self._prev_lat is None:
            self._prev_lat = lat
            return False
        transition = self._prev_lat < 0 and lat >= 0
        self._prev_lat = lat
        return transition

    def predict_next_entry(self, dt: datetime, ground_pos: Tuple) -> Optional[datetime]:
        """
        Прогнозирует следующее вхождение в зону съемки методом последовательных приближений.
        
        Алгоритм:
        1. Постепенно увеличивает время с шагом 60 сек
        2. Для каждого момента проверяет условия съемки
        3. Возвращает первое время, когда условия выполняются
        """
        step = timedelta(seconds=60)
        max_steps = 100  # Максимальное число шагов поиска
        for _ in range(max_steps):
            dt += step
            angles = self.calculate_angles(dt, ground_pos)
            if self.is_in_shooting_zone(angles):
                return dt - step  # Возвращаем начало интервала
        return None

     def find_exact_boundary(self, approx_dt: datetime, 
                           search_window: timedelta, 
                           is_start: bool,
                           ground_pos: Tuple) -> datetime:
        """
        Бинарный поиск границы зоны съемки с пониженной точностью.
        
        Параметры:
        approx_dt: Приблизительное время границы
        search_window: Временное окно поиска (рекомендуется 10-60 секунд)
        is_start: Флаг поиска начала (True) или конца (False) зоны
        ground_pos: Координаты наземной станции
        
        Возвращает:
        datetime: Время границы с точностью до 10 миллисекунд
        """
        low = approx_dt - search_window/2
        high = approx_dt + search_window/2
        
        # 10 итераций обеспечивают точность ~10 мс для 60-секундного окна
        for _ in range(10):
            mid = low + (high - low)/2
            angles = self.calculate_angles(mid, ground_pos)
            in_zone = self.is_in_shooting_zone(angles)
        # Логика бинарного поиска для разных типов границ           
            if (in_zone and is_start) or (not in_zone and not is_start):
                high = mid
            else:
                low = mid
                
        return mid

    def calculate_angles(self, dt: datetime, ground_pos: Tuple) -> Dict[str, float]:
        """Вычисляет геометрические параметры для заданного времени."""
        R_s, V_s = self.get_inertial_position(dt)
        pos_it, _ = get_xyzv_from_latlon(dt, *ground_pos)
        return calculate_angles(R_s, V_s, pos_it)

    def is_in_shooting_zone(self, angles: Dict[str, float]) -> bool:
        """Проверяет выполнение условий съемки: угол визирования 24-55° и видимость."""
        return 24 <= angles['y_grad'] <= 55 and angles['R_0'] < angles['R_e']

def worker(args):
    """Функция для отдельного процесса"""
    tracker, params, time_range = args
    return process_time_range(tracker, params, time_range)

def parallel_calculation(tracker, params, chunks):
    """Распараллеливание с использованием пула процессов"""
    with Pool() as pool:
        args = [(tracker, params, chunk) for chunk in chunks]
        results = pool.map(worker, args)
    return results

def calculate_angles(R_s: Tuple[float, float, float], 
                    V_s: Tuple[float, float, float],
                    pos_it: Tuple[float, float, float]) -> Dict[str, float]:
    """
    Вычисляет углы и расстояния между спутником и наземной точкой.
    
    Формулы основаны на сферической геометрии и законе косинусов.
    
    Параметры:
    R_s - координаты спутника в инерциальной СК (X, Y, Z) [км]
    V_s - вектор скорости спутника [км/с]
    pos_it - координаты наземной точки в инерциальной СК [км]
    
    Возвращает словарь с:
    - R_s: Расстояние от центра Земли до спутника
    - R_e: Расстояние от центра Земли до наземной точки
    - R_0: Дистанция спутник-точка
    - Углы визирования (y_grad) и места (ay_grad) в градусах
    """
    X_s, Y_s, Z_s = R_s
    X_t, Y_t, Z_t = pos_it
    
    # Расчет расстояний по теореме Пифагора
    R_s_norm = math.sqrt(X_s**2 + Y_s**2 + Z_s**2)
    R_e_norm = math.sqrt(X_t**2 + Y_t**2 + Z_t**2)
    R_0_norm = math.sqrt((X_s-X_t)**2 + (Y_s-Y_t)**2 + (Z_s-Z_t)**2)
    V_s_norm = math.sqrt(V_s[0]**2 + V_s[1]**2 + V_s[2]**2)
    
    # Угол визирования (между векторами R_s и R_0)
    y = math.acos((R_0_norm**2 + R_s_norm**2 - R_e_norm**2) / (2 * R_0_norm * R_s_norm))
    
    # Угол места (возвышения антенны)
    ay = math.acos((R_0_norm * math.sin(y)) / R_e_norm)
    
    return {
        'R_s': R_s_norm,
        'R_e': R_e_norm,
        'R_0': R_0_norm,
        'V_s': V_s_norm,
        'y_grad': math.degrees(y),
        'ay_grad': math.degrees(ay)
    }
def calculate_angles_batch(times, ground_pos):
    """Векторизованный расчет для массива временных меток"""
    # Конвертируем входные данные в numpy массивы
    times_np = np.array([t.timestamp() for t in times])
    
    # Здесь должна быть реализация векторных операций
    # Например:
    positions = np.vectorize(lambda t: tracker.get_inertial_position(t))(times_np)
    
    # Возвращаем результаты для всех временных меток
    return [calculate_angles(*pos) for pos in positions]

def split_into_chunks(params, chunk_size=timedelta(minutes=30)):
    """Разбиваем временной диапазон на чанки"""
    chunks = []
    current = params.dt_start
    while current < params.dt_end:
        end = min(current + chunk_size, params.dt_end)
        chunks.append((current, end))
        current = end
    return chunks
def create_orbital_track(tracker: SatelliteTracker,
                        params: ObservationParameters,
                        track_shape: Optional[shapefile.Writer] = None,
                        angle_param: float = 0) -> List[Tuple]:
    """
    Основная функция построения орбитального трека с учетом условий съемки.
    
    Алгоритм работы:
    1. Инициализация параметров цикла
    2. Последовательный перебор временных интервалов
    3. Проверка условий съемки для каждого момента времени
    4. Обработка смены витков
    5. Запись результатов в shapefile
    6. Постобработка для уточнения временных меток
    
    Возвращает список витков с уточненными временными границами.
    """
    lat_t, lon_t, alt_t = params.pos_gt
    t_semki = []  # Список для хранения данных по виткам
    current_orbit = 0  # Текущий номер витка
    orbit_start_time = None  # Время начала текущего витка
    orbit_points = []  # Точки съемки в текущем витке
    
    dt = params.dt_start  # Текущее время моделирования
    computation_delay = timedelta(milliseconds=50)  # Компенсация вычислительной задержки
    
    # Основной цикл моделирования
    while dt < params.dt_end:
        actual_dt = dt - computation_delay  # Коррекция времени
        
        # Расчет углов и расстояний
        angles = tracker.calculate_angles(actual_dt, params.pos_gt)
        
        if tracker.is_in_shooting_zone(angles):
            # Получение текущих параметров спутника
            lon_s, lat_s, alt_s = tracker.get_geodetic_position(actual_dt)
            R_s, V_s = tracker.get_inertial_position(actual_dt)
            
            # Расчет доплеровской частоты
            Fd = calc_f_doplera(
                angle_param, WAVELENGTH, math.radians(angles['ay_grad']), 
                R_s, V_s, angles['R_0'], angles['R_e'],
                angles['R_s'], angles['V_s']
            )

            # Определение смены витка
            is_new_orbit = tracker.check_orbit_transition(actual_dt)
            if is_new_orbit:
                current_orbit += 1

            # Обработка перехода на новый виток
            if is_new_orbit and orbit_start_time is not None:
                t_semki.append((
                    current_orbit - 1,
                    orbit_points.copy(),
                    orbit_start_time,
                    orbit_points[-1][0] if orbit_points else actual_dt,
                    (orbit_points[-1][0] - orbit_start_time) if orbit_points else timedelta(0)
                ))
                orbit_start_time = actual_dt
                orbit_points = []

            # Добавление точки съемки
            orbit_points.append((actual_dt, (lon_s, lat_s, alt_s), Fd))
            
            # Запись в GIS shapefile
            if track_shape:
                track_shape.point(lon_s, lat_s)
                track_shape.record(
                    current_orbit,
                    actual_dt.strftime("%Y-%m-%d %H:%M:%S.%f"),
                    lon_s, lat_s,
                    float(angles['R_s']),
                    float(angles['R_e']),
                    float(angles['R_0']),
                    float(angles['y_grad']),
                    float(angles['ay_grad']),
                    float(WAVELENGTH),
                    float(Fd)
                )

            # Увеличение времени с малым шагом во время съемки
            dt += params.delta_obn
        else:
            # Пропуск неактивных интервалов с прогнозированием
            next_entry = tracker.predict_next_entry(dt, params.pos_gt)
            if next_entry and next_entry < params.dt_end:
                dt = next_entry
            else:
                dt += params.delta

    # Постобработка для уточнения временных границ
    refined_t_semki = []
    for orbit in t_semki:
        # Уточнение начала и конца витка с бинарным поиском
        exact_start = tracker.find_exact_boundary(
            orbit[2], timedelta(seconds=10), True, params.pos_gt)
        exact_end = tracker.find_exact_boundary(
            orbit[3], timedelta(seconds=10), False, params.pos_gt)
        
        refined_t_semki.append((
            orbit[0],
            orbit[1],
            exact_start,
            exact_end,
            exact_end - exact_start
        ))

    # Добавление последнего витка, если данные есть
    if orbit_points:
        exact_start = tracker.find_exact_boundary(
            orbit_start_time, timedelta(seconds=10), True, params.pos_gt)
        exact_end = tracker.find_exact_boundary(
            orbit_points[-1][0], timedelta(seconds=10), False, params.pos_gt)
        
        refined_t_semki.append((
            current_orbit,
            orbit_points,
            exact_start,
            exact_end,
            exact_end - exact_start
        ))

    return refined_t_semki

def send_telegram_notification(message: str):
    """Отправка уведомления в Telegram"""
    try:
        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {
            'chat_id': TELEGRAM_CHAT_ID,
            'text': message,
            'parse_mode': 'HTML'
        }
        requests.post(url, data=payload, timeout=10)
    except Exception as e:
        print(f"Ошибка отправки в Telegram: {e}")

def process_single_orbit(tracker, params, orbit_num, start_time, end_time):
    """Обработка одного витка"""
    orbit_data = []
    dt = start_time
    while dt < end_time:
        # ... (основная логика расчета для одного витка)
        dt += params.delta
    return orbit_data

def create_orbital_track_parallel(tracker, params, max_workers=4):
    """Параллельная версия основной функции"""
    # Сначала находим все временные границы витков
    orbits = detect_all_orbits(tracker, params)
    
    # Параллельная обработка
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = []
        for orbit_num, (start, end) in enumerate(orbits):
            futures.append(executor.submit(
                process_single_orbit,
                tracker, params, orbit_num, start, end
            ))
        
        # Сбор результатов
        results = []
        for future in as_completed(futures):
            results.extend(future.result())
    
    return sorted(results, key=lambda x: x[0])  # Сортировка по времени


def save_to_excel(data: List[Tuple], filename: str, params, tracker: SatelliteTracker):
    """
    Экспорт результатов в формат Excel с автоматической настройкой ширины столбцов.
    
    Структура отчета:
    - Номер витка
    - Временная метка с микросекундами
    - Географические координаты
    - Радиолокационные параметры
    - Геометрические параметры
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Satellite Data"
    
    # Формирование заголовков
    headers = [
        "Виток", "Время", "Долгота", "Широта", "Высота",
        "Fd (Гц)", "R_s (км)", "R_e (км)", "R_0 (км)",
        "Угол визирования (°)", "Угол места (°)"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2  # Начальная строка данных
    for orbit in data:
        orbit_num, points, start_time, end_time, duration = orbit
        for point in points:
            dt, (lon, lat, alt), Fd = point
            # Получение актуальных углов для текущей временной метки
            angles = tracker.calculate_angles(dt, params.pos_gt)
            
            # Запись данных в строку
            ws.cell(row=row, column=1, value=orbit_num)
            ws.cell(row=row, column=2, value=dt.strftime("%Y-%m-%d %H:%M:%S.%f"))  # Формат с микросекундами
            ws.cell(row=row, column=3, value=lon)
            ws.cell(row=row, column=4, value=lat)
            ws.cell(row=row, column=5, value=alt)
            ws.cell(row=row, column=6, value=Fd)
            ws.cell(row=row, column=7, value=angles['R_s'])
            ws.cell(row=row, column=8, value=angles['R_e'])
            ws.cell(row=row, column=9, value=angles['R_0'])
            ws.cell(row=row, column=10, value=angles['y_grad'])
            ws.cell(row=row, column=11, value=angles['ay_grad'])
            
            row += 1  # Переход к следующей строке
    
    # Автоподбор ширины столбцов на основе содержимого
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получение буквенного обозначения столбца
        for cell in col:
            try:
                # Поиск максимальной длины значения в столбце
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Расчет ширины с запасом
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filename)
    print(f"Данные сохранены в Excel файл: {filename}")

def main():
    print("Начало расчета орбит")
    send_telegram_notification("🚀 Запуск расчета орбит")
    
    # Инициализация параметров
    """Функция тестирования основных возможностей модуля"""
    # Инициализация параметров спутника из локального файла
    s_name, tle_1, tle_2 = read_tle_base_file(56756)  # Чтение TLE для спутника с номером 56756
    
    # Координаты наземной станции (Санкт-Петербург)
    pos_gt = (59.95, 30.316667, 0)
    
    # Настройка параметров наблюдения
    params = ObservationParameters()
    params.pos_gt = pos_gt
    params.dt_start = datetime(2024, 2, 21, 3, 0, 0)  # Начало наблюдения
    params.delta = timedelta(seconds=10)  # Основной шаг расчета
    params.delta_obn = timedelta(seconds=5)  # Шаг во время съемки
    params.dt_end = params.dt_start + timedelta(days=2)  # Окончание через 2 дня
    
    # Инициализация трекера с загруженными TLE-данными
    tracker = SatelliteTracker(tle_1, tle_2)
    
 
    # Параллельный расчет для каждого угла
    for angle in ANGLES_TO_TEST:
        start_time = time.time()
        print(f"\nРасчет для угла {angle}°")
        send_telegram_notification(f"🔽 Начало расчета для угла {angle}°")
        
        results = parallel_orbit_calculation(tracker, params)
        
        # Сохранение результатов
        excel_filename = f"result/orbit_angle_{angle}.xlsx"
        save_to_excel(results, excel_filename)
        
        duration = time.time() - start_time
        print(f"Угол {angle}°: расчет занял {duration:.2f} сек")
        send_telegram_notification(
            f"🔼 Завершен угол {angle}°\n"
            f"⏱ Время: {duration:.2f} сек\n"
            f"📊 Точек: {len(results)}"
        )
    
    print("\nВсе расчеты завершены!")
    send_telegram_notification("✅ Все расчеты успешно завершены")

if __name__ == "__main__":
    main()