"""
Модуль для расчета орбитальных параметров спутника и условий съемки наземных объектов с использованием параллельных вычислений
"""

# Импорт необходимых библиотек
import math
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple
import concurrent.futures
import multiprocessing

import shapefile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pyorbital.orbital import Orbital

from calc_cord import get_xyzv_from_latlon
from calc_F_L import calc_f_doplera, calc_lamda
from read_TBF import read_tle_base_file, read_tle_base_internet

# Глобальные константы
EARTH_RADIUS = 6378.140
EARTH_ROTATION_RATE = 7.2292115E-5
WAVELENGTH = 0.000096
ORBITAL_PERIOD = 5689

class ObservationParameters:
    def __init__(self):
        self.dt_start = None
        self.dt_end = None
        self.delta = None
        self.delta_obn = None
        self.pos_gt = None

class SatelliteTracker:
    def __init__(self, tle_1: str, tle_2: str):
        self.orb = Orbital("N", line1=tle_1, line2=tle_2)
        self.tle_1 = tle_1
        self.tle_2 = tle_2

    def get_geodetic_position(self, utc_time: datetime) -> Tuple[float, float, float]:
        return self.orb.get_lonlatalt(utc_time)

    def get_inertial_position(self, utc_time: datetime) -> Tuple[Tuple[float, float, float], Tuple[float, float, float]]:
        return self.orb.get_position(utc_time, False)

    def check_orbit_transition(self, prev_lat: Optional[float], dt: datetime) -> Tuple[bool, float]:
        """Определяет переход витка и возвращает текущую широту."""
        _, lat, _ = self.get_geodetic_position(dt)
        if prev_lat is None:
            return False, lat
        return prev_lat < 0 and lat >= 0, lat

    def find_exact_boundary(self, approx_dt: datetime, search_window: timedelta, is_start: bool, ground_pos: Tuple) -> datetime:
        low = approx_dt - search_window // 2
        high = approx_dt + search_window // 2
        
        for _ in range(20):
            mid = low + (high - low) / 2
            angles = self.calculate_angles(mid, ground_pos)
            in_zone = self.is_in_shooting_zone(angles, 0)  # angle_param временно 0
            if (in_zone and is_start) or (not in_zone and not is_start):
                high = mid
            else:
                low = mid
        return mid

    def calculate_angles(self, dt: datetime, ground_pos: Tuple) -> Dict[str, float]:
        R_s, V_s = self.get_inertial_position(dt)
        pos_it, _ = get_xyzv_from_latlon(dt, *ground_pos)
        return calculate_angles(R_s, V_s, pos_it)

    def is_in_shooting_zone(self, angles: Dict[str, float], angle_param: float) -> bool:
        """Проверка нахождения в зоне съемки с учетом angle_param."""
        return (
            24 <= angles['y_grad'] <= 55 
            and angles['R_0'] < angles['R_e']
            and abs(angles['ay_grad'] - angle_param) <= 2  # Допуск ±2°
        )

def calculate_angles(R_s: Tuple[float, float, float], 
                    V_s: Tuple[float, float, float],
                    pos_it: Tuple[float, float, float]) -> Dict[str, float]:
    X_s, Y_s, Z_s = R_s
    X_t, Y_t, Z_t = pos_it
    
    R_s_norm = math.sqrt(X_s**2 + Y_s**2 + Z_s**2)
    R_e_norm = math.sqrt(X_t**2 + Y_t**2 + Z_t**2)
    R_0_norm = math.sqrt((X_s - X_t)**2 + (Y_s - Y_t)**2 + (Z_s - Z_t)**2)
    V_s_norm = math.sqrt(V_s[0]**2 + V_s[1]**2 + V_s[2]**2)
    
    y = math.acos((R_0_norm**2 + R_s_norm**2 - R_e_norm**2) / (2 * R_0_norm * R_s_norm))
    ay = math.acos((R_0_norm * math.sin(y)) / R_e_norm)
    
    return {
        'R_s': R_s_norm,
        'R_e': R_e_norm,
        'R_0': R_0_norm,
        'V_s': V_s_norm,
        'y_grad': math.degrees(y),
        'ay_grad': math.degrees(ay)
    }

def find_orbit_transitions(tracker: SatelliteTracker, start_dt: datetime, end_dt: datetime) -> List[datetime]:
    transitions = []
    current_dt = start_dt
    prev_lat = None
    step = timedelta(seconds=60)
    
    while current_dt < end_dt:
        transition, prev_lat = tracker.check_orbit_transition(prev_lat, current_dt)
        if transition:
            low = current_dt - step
            high = current_dt
            for _ in range(10):  # Увеличено количество итераций
                mid = low + (high - low) / 2
                mid_lat = tracker.get_geodetic_position(mid)[1]
                if mid_lat >= 0:
                    high = mid
                else:
                    low = mid
            transitions.append(high)
            current_dt = high
        current_dt += step
    return transitions

def process_time_step(tle_1: str, tle_2: str, dt: datetime, ground_pos: Tuple, angle_param: float) -> dict:
    tracker = SatelliteTracker(tle_1, tle_2)
    angles = tracker.calculate_angles(dt, ground_pos)
    in_zone = tracker.is_in_shooting_zone(angles, angle_param)  # Учет angle_param
    lon_s, lat_s, alt_s = tracker.get_geodetic_position(dt)
    R_s, V_s = tracker.get_inertial_position(dt)
    X_s, Y_s, Z_s = R_s
    X_t, Y_t, Z_t = get_xyzv_from_latlon(dt, *ground_pos)[0]
    
    Fd = calc_f_doplera(
        angle_param, WAVELENGTH, math.radians(angles['ay_grad']),
        (X_s, Y_s, Z_s), V_s,
        angles['R_0'], angles['R_e'],
        angles['R_s'], angles['V_s']
    )
    
    return {
        'dt': dt,
        'in_zone': in_zone,
        'pos': (lon_s, lat_s, alt_s),
        'Fd': Fd,
        'angles': angles
    }

def create_orbital_track(tracker: SatelliteTracker,
                        params: ObservationParameters,
                        track_shape: Optional[shapefile.Writer] = None,
                        angle_param: float = 0) -> List[Tuple]:
    transitions = find_orbit_transitions(tracker, params.dt_start, params.dt_end)
    orbit_boundaries = [params.dt_start] + transitions + [params.dt_end]
    orbits = [(i, start, end) for i, (start, end) in enumerate(zip(orbit_boundaries[:-1], orbit_boundaries[1:]))]

    all_results = []
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = []
        for orbit_num, start, end in orbits:
            current_dt = start
            while current_dt < end:
                futures.append(executor.submit(
                    process_time_step,
                    tracker.tle_1,
                    tracker.tle_2,
                    current_dt,
                    params.pos_gt,
                    angle_param
                ))
                current_dt += params.delta_obn
        
        for future in concurrent.futures.as_completed(futures):
            all_results.append(future.result())

    all_results.sort(key=lambda x: x['dt'])
    t_semki = []
    current_orbit = 0
    orbit_points = []
    
    for res in all_results:
        if res['in_zone']:
            orbit_points.append((res['dt'], res['pos'], res['Fd']))
        else:
            if orbit_points:
                exact_start = tracker.find_exact_boundary(
                    orbit_points[0][0], params.delta_obn, True, params.pos_gt)
                exact_end = tracker.find_exact_boundary(
                    orbit_points[-1][0], params.delta_obn, False, params.pos_gt)
                t_semki.append((
                    current_orbit,
                    orbit_points,
                    exact_start,
                    exact_end,
                    exact_end - exact_start
                ))
                current_orbit += 1
                orbit_points = []
    
    if orbit_points:
        exact_start = tracker.find_exact_boundary(
            orbit_points[0][0], params.delta_obn, True, params.pos_gt)
        exact_end = tracker.find_exact_boundary(
            orbit_points[-1][0], params.delta_obn, False, params.pos_gt)
        t_semki.append((
            current_orbit,
            orbit_points,
            exact_start,
            exact_end,
            exact_end - exact_start
        ))
    
    if track_shape:
        for orbit in t_semki:
            for point in orbit[1]:
                dt, pos, Fd = point
                lon_s, lat_s, alt_s = pos
                angles = tracker.calculate_angles(dt, params.pos_gt)
                track_shape.point(lon_s, lat_s)
                track_shape.record(
                    orbit[0],
                    dt.strftime("%Y-%m-%d %H:%M:%S.%f"),
                    lon_s, lat_s,
                    angles['R_s'],
                    angles['R_e'],
                    angles['R_0'],
                    angles['y_grad'],
                    angles['ay_grad'],
                    WAVELENGTH,
                    Fd
                )
    
    return t_semki

def save_to_excel(data: List[Tuple], filename: str, params, tracker: SatelliteTracker):
    wb = Workbook()
    ws = wb.active
    ws.title = "Satellite Data"
    
    headers = [
        "Виток", "Время", "Долгота", "Широта", "Высота",
        "Fd (Гц)", "R_s (км)", "R_e (км)", "R_0 (км)",
        "Угол визирования (°)", "Угол места (°)"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2
    for orbit in data:
        orbit_num, points, start_time, end_time, duration = orbit
        for point in points:
            dt, (lon, lat, alt), Fd = point
            angles = tracker.calculate_angles(dt, params.pos_gt)
            ws.cell(row=row, column=1, value=orbit_num)
            ws.cell(row=row, column=2, value=dt.strftime("%Y-%m-%d %H:%M:%S.%f"))
            ws.cell(row=row, column=3, value=lon)
            ws.cell(row=row, column=4, value=lat)
            ws.cell(row=row, column=5, value=alt)
            ws.cell(row=row, column=6, value=Fd)
            ws.cell(row=row, column=7, value=angles['R_s'])
            ws.cell(row=row, column=8, value=angles['R_e'])
            ws.cell(row=row, column=9, value=angles['R_0'])
            ws.cell(row=row, column=10, value=angles['y_grad'])
            ws.cell(row=row, column=11, value=angles['ay_grad'])
            row += 1
    
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(filename)
    print(f"Данные сохранены в Excel файл: {filename}")

def _test():
    s_name, tle_1, tle_2 = read_tle_base_file(56756)
    pos_gt = (59.95, 30.316667, 0)
    
    params = ObservationParameters()
    params.pos_gt = pos_gt
    params.dt_start = datetime(2024, 2, 21, 3, 0, 0)
    params.delta = timedelta(seconds=10)
    params.delta_obn = timedelta(seconds=5)
    params.dt_end = params.dt_start + timedelta(days=2)
    
    tracker = SatelliteTracker(tle_1, tle_2)
    
    shp_filename = f"result/{s_name}"
    track_shape = shapefile.Writer(shp_filename, shapefile.POINT)
    
    try:
        fields = [
            ("ID", "N", 10),
            ("TIME", "C", 40),
            ("LON", "F", 10, 10),
            ("LAT", "F", 10, 10),
            ("R_s", "F", 10, 5),
            ("R_t", "F", 10, 5),
            ("R_n", "F", 10, 5),
            ("ϒ", "F", 10, 5),
            ("φ", "F", 10, 5),
            ("λ", "F", 10, 10),
            ("f", "F", 10, 5)
        ]
        for field in fields:
            track_shape.field(*field if len(field) == 3 else field[:4])
        
        angle_values = range(88, 92, 1)
        results = {}
        
        for angle in angle_values:
            t_semki = create_orbital_track(tracker, params, track_shape, angle)
            results[angle] = t_semki
            
            excel_filename = f"result/{s_name}_angle_{angle}.xlsx"
            save_to_excel(t_semki, excel_filename, params, tracker)
            
            print(f"\nУгол {angle}°:")
            for orbit in t_semki:
                fd_values = [point[2] for point in orbit[1]]
                if fd_values:
                    print(f"Виток {orbit[0]}: точек {len(orbit[1])}, Fd от {min(fd_values):.2f} до {max(fd_values):.2f} Гц")
        
        track_shape.save(shp_filename)
        print(f"\nShapefile сохранен: {shp_filename}.shp")
        
    except Exception as e:
        print(f"Ошибка: {str(e)}")
    finally:
        if 'track_shape' in locals():
            try:
                track_shape.close()
            except:
                pass
    
    try:
        with open(f"{shp_filename}.prj", "w") as prj:
            prj.write('GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]')
    except Exception as e:
        print(f"Ошибка при создании PRJ-файла: {e}")

if __name__ == "__main__":
    multiprocessing.freeze_support()
    _test()