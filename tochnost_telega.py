"""
Модуль для расчета орбитальных параметров спутника и условий съемки наземных объектов
"""

# Импорт необходимых библиотек
import math
import os
import zipfile
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import requests
import shapefile
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pyorbital.orbital import Orbital

# Локальные модули
from calc_cord import get_xyzv_from_latlon
from calc_F_L import calc_f_doplera, calc_lamda
from read_TBF import read_tle_base_file, read_tle_base_internet

# ----------------------------
# Настройка логирования
# ----------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Загрузка переменных окружения
load_dotenv()

# ----------------------------
# Глобальные константы
# ----------------------------
EARTH_RADIUS = 6378.140  # км
EARTH_ROTATION_RATE = 7.2292115E-5  # рад/с
WAVELENGTH = 0.000096  # метра (X-диапазон)
ORBITAL_PERIOD = 5689  # секунд (~94 мин)
RESULT_DIR = "result"
ZIP_NAME = "result.zip"

class ObservationParameters:
    """Контейнер для параметров наблюдения"""
    def __init__(self):
        self.dt_start = None
        self.dt_end = None
        self.delta = None
        self.delta_obn = None
        self.pos_gt = None

class SatelliteTracker:
    """Орбитальный калькулятор на основе TLE"""
    
    def __init__(self, tle_1: str, tle_2: str):
        self.orb = Orbital("N", line1=tle_1, line2=tle_2)
        self.tle_1 = tle_1
        self.tle_2 = tle_2
        self._prev_lat = None

    def get_geodetic_position(self, utc_time: datetime) -> Tuple[float, float, float]:
        return self.orb.get_lonlatalt(utc_time)

    def get_inertial_position(self, utc_time: datetime) -> Tuple[Tuple[float, float, float], Tuple[float, float, float]]:
        return self.orb.get_position(utc_time, False)

    def check_orbit_transition(self, dt: datetime) -> bool:
        _, lat, _ = self.get_geodetic_position(dt)
        if self._prev_lat is None:
            self._prev_lat = lat
            return False
        transition = self._prev_lat < 0 and lat >= 0
        self._prev_lat = lat
        return transition

    def predict_next_entry(self, dt: datetime, ground_pos: Tuple) -> Optional[datetime]:
        step = timedelta(seconds=60)
        for _ in range(100):
            dt += step
            angles = self.calculate_angles(dt, ground_pos)
            if self.is_in_shooting_zone(angles):
                return dt - step
        return None

    def find_exact_boundary(self, approx_dt: datetime, 
                           search_window: timedelta, 
                           is_start: bool,
                           ground_pos: Tuple) -> datetime:
        low = approx_dt - search_window/2
        high = approx_dt + search_window/2
        
        for _ in range(10):
            mid = low + (high - low)/2
            angles = self.calculate_angles(mid, ground_pos)
            in_zone = self.is_in_shooting_zone(angles)
            
            if (in_zone and is_start) or (not in_zone and not is_start):
                high = mid
            else:
                low = mid
        return mid

    def calculate_angles(self, dt: datetime, ground_pos: Tuple) -> Dict[str, float]:
        R_s, V_s = self.get_inertial_position(dt)
        pos_it, _ = get_xyzv_from_latlon(dt, *ground_pos)
        return calculate_angles(R_s, V_s, pos_it)

    def is_in_shooting_zone(self, angles: Dict[str, float]) -> bool:
        return 24 <= angles['y_grad'] <= 55 and angles['R_0'] < angles['R_e']

def create_zip_archive() -> bool:
    """Создает zip-архив с результатами"""
    try:
        with zipfile.ZipFile(ZIP_NAME, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(RESULT_DIR):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, start=RESULT_DIR)
                    zipf.write(file_path, arcname)
        logger.info(f"Архив {ZIP_NAME} успешно создан")
        return True
    except Exception as e:
        logger.error(f"Ошибка создания архива: {str(e)}")
        return False

def send_to_telegram() -> bool:
    """Отправляет архив через Telegram бота"""
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    
    if not bot_token or not chat_id:
        logger.error("Не заданы учетные данные Telegram")
        return False

    url = f"https://api.telegram.org/bot{bot_token}/sendDocument"
    
    try:
        with open(ZIP_NAME, 'rb') as file:
            response = requests.post(
                url,
                files={'document': file},
                data={'chat_id': chat_id},
                timeout=30
            )
            if response.status_code == 200:
                logger.info("Архив успешно отправлен")
                return True
            logger.error(f"Ошибка отправки: {response.text}")
            return False
    except Exception as e:
        logger.error(f"Ошибка сети: {str(e)}")
        return False

def calculate_angles(R_s: Tuple[float, float, float], 
                    V_s: Tuple[float, float, float],
                    pos_it: Tuple[float, float, float]) -> Dict[str, float]:
    X_s, Y_s, Z_s = R_s
    X_t, Y_t, Z_t = pos_it
    
    R_s_norm = math.sqrt(sum(x**2 for x in R_s))
    R_e_norm = math.sqrt(sum(x**2 for x in pos_it))
    R_0_norm = math.sqrt(sum((s - t)**2 for s, t in zip(R_s, pos_it)))
    V_s_norm = math.sqrt(sum(v**2 for v in V_s))
    
    y = math.acos((R_0_norm**2 + R_s_norm**2 - R_e_norm**2) / (2 * R_0_norm * R_s_norm))
    ay = math.acos((R_0_norm * math.sin(y)) / R_e_norm)
    
    return {
        'R_s': R_s_norm,
        'R_e': R_e_norm,
        'R_0': R_0_norm,
        'V_s': V_s_norm,
        'y_grad': math.degrees(y),
        'ay_grad': math.degrees(ay)
    }

def create_orbital_track(tracker: SatelliteTracker,
                        params: ObservationParameters,
                        track_shape: Optional[shapefile.Writer] = None,
                        angle_param: float = 0) -> List[Tuple]:
    
    lat_t, lon_t, alt_t = params.pos_gt
    t_semki = []
    current_orbit = 0
    orbit_start_time = None
    orbit_points = []
    
    dt = params.dt_start
    computation_delay = timedelta(milliseconds=50)
    
    while dt < params.dt_end:
        actual_dt = dt - computation_delay
        
        angles = tracker.calculate_angles(actual_dt, params.pos_gt)
        
        if tracker.is_in_shooting_zone(angles):
            lon_s, lat_s, alt_s = tracker.get_geodetic_position(actual_dt)
            R_s, V_s = tracker.get_inertial_position(actual_dt)
            
            Fd = calc_f_doplera(
                angle_param, WAVELENGTH, math.radians(angles['ay_grad']), 
                R_s, V_s, angles['R_0'], angles['R_e'],
                angles['R_s'], angles['V_s']
            )

            is_new_orbit = tracker.check_orbit_transition(actual_dt)
            if is_new_orbit:
                current_orbit += 1

            if is_new_orbit and orbit_start_time is not None:
                t_semki.append((
                    current_orbit - 1,
                    orbit_points.copy(),
                    orbit_start_time,
                    orbit_points[-1][0] if orbit_points else actual_dt,
                    (orbit_points[-1][0] - orbit_start_time) if orbit_points else timedelta(0)
                ))
                orbit_start_time = actual_dt
                orbit_points = []

            orbit_points.append((actual_dt, (lon_s, lat_s, alt_s), Fd))
            
            if track_shape:
                track_shape.point(lon_s, lat_s)
                track_shape.record(
                    current_orbit,
                    actual_dt.strftime("%Y-%m-%d %H:%M:%S.%f"),
                    lon_s, lat_s,
                    *map(float, [angles[k] for k in ['R_s', 'R_e', 'R_0', 'y_grad', 'ay_grad']]),
                    float(WAVELENGTH),
                    float(Fd)
                )

            dt += params.delta_obn
        else:
            next_entry = tracker.predict_next_entry(dt, params.pos_gt)
            dt = next_entry if next_entry and next_entry < params.dt_end else dt + params.delta

    # Постобработка данных
    refined_t_semki = []
    for orbit in t_semki:
        exact_start = tracker.find_exact_boundary(
            orbit[2], timedelta(seconds=10), True, params.pos_gt)
        exact_end = tracker.find_exact_boundary(
            orbit[3], timedelta(seconds=10), False, params.pos_gt)
        
        refined_t_semki.append((
            orbit[0],
            orbit[1],
            exact_start,
            exact_end,
            exact_end - exact_start
        ))

    if orbit_points:
        exact_start = tracker.find_exact_boundary(
            orbit_start_time, timedelta(seconds=10), True, params.pos_gt)
        exact_end = tracker.find_exact_boundary(
            orbit_points[-1][0], timedelta(seconds=10), False, params.pos_gt)
        
        refined_t_semki.append((
            current_orbit,
            orbit_points,
            exact_start,
            exact_end,
            exact_end - exact_start
        ))

    return refined_t_semki

def save_to_excel(data: List[Tuple], filename: str, params, tracker: SatelliteTracker):
    wb = Workbook()
    ws = wb.active
    ws.title = "Satellite Data"
    
    headers = [
        "Виток", "Время", "Долгота", "Широта", "Высота",
        "Fd (Гц)", "R_s (км)", "R_e (км)", "R_0 (км)",
        "Угол визирования (°)", "Угол места (°)"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    row = 2
    for orbit in data:
        orbit_num, points, *_ = orbit
        for point in points:
            dt, (lon, lat, alt), Fd = point
            angles = tracker.calculate_angles(dt, params.pos_gt)
            
            ws.cell(row=row, column=1, value=orbit_num)
            ws.cell(row=row, column=2, value=dt.strftime("%Y-%m-%d %H:%M:%S.%f"))
            for i, val in enumerate([lon, lat, alt, Fd], 3):
                ws.cell(row=row, column=i, value=val)
            for i, key in enumerate(['R_s', 'R_e', 'R_0', 'y_grad', 'ay_grad'], 7):
                ws.cell(row=row, column=i, value=angles[key])
            row += 1
    
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = (max_len + 2) * 1.2
    
    wb.save(filename)
    logger.info(f"Excel-файл сохранен: {filename}")

def _test():
    """Основная тестовая функция"""
    try:
        # Инициализация
        s_name, tle_1, tle_2 = read_tle_base_file(56756)
        pos_gt = (59.95, 30.316667, 0)
        
        # Создание директории результатов
        os.makedirs(RESULT_DIR, exist_ok=True)

        # Настройка параметров
        params = ObservationParameters()
        params.pos_gt = pos_gt
        params.dt_start = datetime(2024, 2, 21, 3, 0, 0)
        params.delta = timedelta(seconds=10)
        params.delta_obn = timedelta(seconds=5)
        params.dt_end = params.dt_start + timedelta(days=2)
        
        # Инициализация трекера
        tracker = SatelliteTracker(tle_1, tle_2)
        
        # Настройка shapefile
        shp_filename = f"{RESULT_DIR}/{s_name}"
        with shapefile.Writer(shp_filename, shapefile.POINT) as track_shape:
            # Определение полей
            fields = [
                ("ID", "N", 10), ("TIME", "C", 40),
                ("LON", "F", 10, 10), ("LAT", "F", 10, 10),
                ("R_s", "F", 10, 5), ("R_t", "F", 10, 5),
                ("R_n", "F", 10, 5), ("ϒ", "F", 10, 5),
                ("φ", "F", 10, 5), ("λ", "F", 10, 10),
                ("f", "F", 10, 5)
            ]
            for field in fields:
                track_shape.field(*field if len(field) == 3 else field[:4])
            
            # Расчет данных
            for angle in range(85, 95, 1):
                t_semki = create_orbital_track(tracker, params, track_shape, angle)
                excel_filename = f"{RESULT_DIR}/{s_name}_angle_{angle}.xlsx"
                save_to_excel(t_semki, excel_filename, params, tracker)
                
                logger.info(f"Угол {angle}° обработан")
                for orbit in t_semki:
                    fd_values = [p[2] for p in orbit[1]]
                    if fd_values:
                        logger.info(f"Виток {orbit[0]}: точек {len(orbit[1])}, Fd: {min(fd_values):.2f}-{max(fd_values):.2f} Гц")

            track_shape.save(shp_filename)
            logger.info(f"Shapefile сохранен: {shp_filename}.shp")

        # Создание PRJ-файла
        prj_content = 'GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]'
        with open(f"{shp_filename}.prj", "w") as f:
            f.write(prj_content)

        # Архивирование и отправка
        if create_zip_archive():
            if send_to_telegram():
                logger.info("Процесс завершен успешно")
            else:
                logger.warning("Не удалось отправить архив")
        else:
            logger.error("Не удалось создать архив")

    except Exception as e:
        logger.error(f"Критическая ошибка: {str(e)}", exc_info=True)
    finally:
        if os.path.exists(ZIP_NAME):
            try:
                os.remove(ZIP_NAME)
                logger.info("Временный архив удален")
            except Exception as e:
                logger.error(f"Ошибка удаления архива: {str(e)}")

if __name__ == "__main__":
    _test()